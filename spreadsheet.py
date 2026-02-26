"""
spreadsheet.py — Write extraction results to a color-coded Excel file.
"""

import openpyxl
from openpyxl.styles import PatternFill, Font

COLUMNS = [
    "image_file",
    "filename_integer",
    "datamatrix_integer",
    "datamatrix_match",
    "transcribed_text",
    "transcription_confidence",
    "transcription_comments",
    "sample_number",
    "date",
    "sampling_event",
    "species",
    "tissue",
    "notes",
    "parse_confidence",
    "parse_comments",
]

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF")
RED_FILL = PatternFill("solid", fgColor="FFCCCC")
YELLOW_FILL = PatternFill("solid", fgColor="FFFACC")


def write_spreadsheet(rows: list[dict], output_path: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Vial Labels"

    # Header row
    for col, name in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    for row_idx, row in enumerate(rows, 2):
        for col, name in enumerate(COLUMNS, 1):
            ws.cell(row=row_idx, column=col, value=row.get(name, ""))

        # Color-code rows by minimum confidence
        t_conf = row.get("transcription_confidence") or 10
        p_conf = row.get("parse_confidence") or 10
        min_conf = min(t_conf, p_conf)
        if min_conf <= 3:
            fill = RED_FILL
        elif min_conf <= 6:
            fill = YELLOW_FILL
        else:
            fill = None

        if fill:
            for col in range(1, len(COLUMNS) + 1):
                ws.cell(row=row_idx, column=col).fill = fill

        # Flag DataMatrix mismatches in that cell specifically
        if row.get("datamatrix_match") is False:
            ws.cell(row=row_idx, column=COLUMNS.index("datamatrix_match") + 1).fill = RED_FILL

    # Auto-width columns
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)

    wb.save(output_path)
    print(f"Saved {len(rows)} rows to {output_path}")
